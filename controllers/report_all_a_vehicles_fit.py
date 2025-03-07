import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import os, traceback
from database import VMS_DB

class Report:
    def __init__(self, db_to_display=None, main_heading = None):
        self.current_date = datetime.now().strftime('%d-%m-%Y')
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = os.path.join(downloads_path, f"All_A_Veh_Fitness_report_{timestamp}.xlsx")
        self.db_to_display = db_to_display
        self.main_header = main_heading

    def generate_report(self):
        try:
            db = VMS_DB()
            weapons = db.get_all_a_vehicle_fit(page=None, page_size=None)

            self.groupColors = {
                "Basic Details": "FF5733",  # Red-Orange
                "Cooling Sys": "33A1FF",     # Blue
                "Hyd Ramp": "28B463",    # Green
                "Lub Sys": "F1C40F",     # Yellow
                "Tr Sys": "8E44AD",  # Purple
                "Bty & Assys": "E67E22",  # Orange
                "Boggy Wh": "7F8C8D",         # Gray
                "Brk Sys": "2C3E50",  # Dark Blue
                "Elec Sys": "D35400",  # Dark Orange
                "Air Intake Sys": "C0392B",  # Dark Red
                "Tx Sys": "16A085",  # Teal
                "Steering Con": "A569BD",         # Soft Purple
                "Fuel Sys": "5D6D7E",         # Slate Blue-Gray
                "Creation Details": "008B8B" # Deep Cyan
            }
            
            if not weapons:
                print("No data found to export.")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "A Vehicle Report"

            # Header
            ws.merge_cells("A1:AG1")
            ws["A1"] = "44 AK HAT Battalion, Pakistan Army"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="left")
            
            ws.merge_cells("A2:AG2")
            ws["A2"] = "A Vehicle Maintenance Report"
            ws["A2"].font = Font(size=12, bold=True)
            ws["A2"].alignment = Alignment(horizontal="left")
            
            ws.merge_cells("A3:AG3")
            ws["A3"] = f"Date: {self.current_date}"
            ws["A3"].font = Font(size=10, italic=True)
            ws["A3"].alignment = Alignment(horizontal="left")

            # Leave 2 rows empty
            ws.append([])
            ws.append([])

            # Column Headers
            # Column Headers with Category Headers
            col_index = 1
            for main_heading, sub_columns in self.main_header.items():
                ws.merge_cells(start_row=6, start_column=col_index, end_row=6, end_column=col_index + len(sub_columns) - 1)
                cell = ws.cell(row=6, column=col_index, value=main_heading)
                cell.font = Font(size=14, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=self.groupColors[main_heading], fill_type="solid")
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                col_index += len(sub_columns)
            
            # ws.append([db_to_display[col] for columns in main_header.values() for col in columns])

            all_columns = [db_colum for db_colum, display_col in self.db_to_display.items()]
            for col_num, column_title in enumerate(all_columns, start=1):
                cell = ws.cell(row=7, column=col_num, value=self.db_to_display[column_title])
                cell.font = Font(size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(style='thin'))

            # Adding data
            for weapon in weapons:
                row_data = []
                for db_col in all_columns:
                    row_data.append(weapon.get(db_col, ""))
                ws.append(row_data)
            
            # Now loop over all data rows (data starts at row 8 in this example because rows 1-7 are header)
            for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
                for idx, cell in enumerate(row, start=1):
                    col_key = all_columns[idx - 1]
                    if col_key not in  ['ba_no_input', "make_input", "type_input", "CI_input", "In_Svc_input", "created_by", "created_at"] and cell.value:
                        try:
                            if cell.value in  ["Unsvc", "Incomplete",  "Unsatisfactory", "Down"]:
                                cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
                                cell.font = Font(color="FFFFFF", bold=True)
                        except ValueError:
                            print(f"Invalid date format for {cell.value} in column {col_key}")
                            continue

            # Auto-Adjust Column Widths
            for col in ws.columns:
                max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max_length + 5

            # Save file
            wb.save(self.filename)
            print(f"Report saved at: {self.filename}")
            return True
        
        except Exception as e:
            traceback.print_exc()
            print(f"Exception in generate_report: {e}")
            return False
        
    def date_rules(self, cell, cell_value, no_of_month, no_of_days):
        # print(f"IN Rule: {type(cell_value)}, {cell_value},  {date.today()}")
        difference = relativedelta(date.today(), cell_value)
        months_diff = difference.years * 12 + difference.months
        days_diff = difference.days
        # print(f"{months_diff} : {days_diff}")

        if months_diff >= no_of_month:
            cell.fill = PatternFill(start_color="FF0000", fill_type="solid")  # Red BG
            cell.font = Font(color="FFFFFF", bold=True)  # White text for visibility
        elif months_diff >= (no_of_month-1) and days_diff >= no_of_days and days_diff <= (no_of_days+10):
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")  # Yellow BG
            cell.font = Font(color="000000", bold=True)  # Black text for visibility

# Usage Example
if __name__ == "__main__":
    report = Report()
    report.generate_report()
