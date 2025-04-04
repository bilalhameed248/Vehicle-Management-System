import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import os, traceback
from database import VMS_DB

class Report:
    def __init__(self, db_to_display=None, main_heading = None):
        self.current_date = datetime.now().strftime('%d-%m-%Y')
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = os.path.join(downloads_path, f"all_vehicles_report_{timestamp}.xlsx")
        self.db_to_display = db_to_display
        self.main_header = main_heading

    def generate_report(self):
        try:
            db = VMS_DB()
            vehicles = db.get_all_vehicle(page=None, page_size=None)

            self.groupColors = {
                "Basic Details": "FF5733",  # Red-Orange
                "Oil Filter": "33A1FF",     # Blue
                "Fuel Filter": "28B463",    # Green
                "Air Filter": "F1C40F",     # Yellow
                "Transmission Filter": "8E44AD",  # Purple
                "Differential Oil": "E67E22",  # Orange
                "Battery Info": "2C3E50",  # Dark Blue
                "Flushing Info": "D35400",  # Dark Orange
                "Greasing Info": "C0392B",  # Dark Red
                "General Maint": "16A085",  # Teal
                "Overhaul": "7F8C8D",        # Gray
                "Status & Creation Details": "008B8B" # Deep Cyan
            }
            
            if not vehicles:
                print("No data found to export.")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vehicle Report"

            # Header
            ws.merge_cells("A1:AG1")
            ws["A1"] = "44 AK HAT Battalion, Pakistan Army"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="left")
            
            ws.merge_cells("A2:AG2")
            ws["A2"] = "Vehicle Maintenance Report"
            ws["A2"].font = Font(size=12, bold=True)
            ws["A2"].alignment = Alignment(horizontal="left")
            
            ws.merge_cells("A3:AG3")
            ws["A3"] = f"Date: {self.current_date}"
            ws["A3"].font = Font(size=10, italic=True)
            ws["A3"].alignment = Alignment(horizontal="left")

            # Leave 2 rows empty
            ws.append([])
            ws.append([])

            # Column Headers with Category Headers
            col_index = 1
            for main_heading, sub_columns in self.main_header.items():
                ws.merge_cells(start_row=6, start_column=col_index, end_row=6, end_column=col_index + len(sub_columns) - 1)
                cell = ws.cell(row=6, column=col_index, value=main_heading)
                cell.font = Font(size=14, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=self.groupColors[main_heading], fill_type="solid")
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                col_index += len(sub_columns)
            

            # Style Column Headers 
            all_columns = [db_colum for db_colum, display_col in self.db_to_display.items()]
            for col_num, column_title in enumerate(all_columns, start=1):
                cell = ws.cell(row=7, column=col_num, value=self.db_to_display[column_title])
                cell.font = Font(size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(style='thin'))

            # Adding data
            for vehicle in vehicles:
                row_data = []
                for db_col in all_columns:
                    row_data.append(vehicle.get(db_col, ""))
                ws.append(row_data)

            issue_due_mapping = {
                'due_date_oil_filter': 6,
                'due_date_fuel_filter': 12,
                'due_date_air_filter': 18,
                'due_date_transmission_filter': 18,
                'due_date_differential_oil': 18,
                'battery_due_date': 42,
                'flusing_due_date': 4,
                'greasing_due_date': 3,
            }

            # Now loop over all data rows (data starts at row 8 in this example because rows 1-7 are header)
            for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
                for idx, cell in enumerate(row, start=1):
                    col_key = all_columns[idx - 1]
                    if col_key in issue_due_mapping and cell.value:
                        try:
                            previous_cell_value = row[idx - 2].value if idx > 1 else None 
                            date_obj = datetime.strptime(str(previous_cell_value), "%Y-%m-%d").date()
                            self.date_rules(cell, date_obj, issue_due_mapping[col_key], 20)
                        except ValueError:
                            print(f"Invalid date format for {cell.value} in column {col_key}")
                            continue

            # Auto-Adjust Column Widths
            for col in ws.columns:
                max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max_length + 5

            # Save file
            wb.save(self.filename)
            print(f"Report saved at: {self.filename}")
            return True
        
        except Exception as e:
            traceback.print_exc()
            print(f"Exception in generate_report: {e}")
            return False
        
    def date_rules(self, cell, cell_value, no_of_month, no_of_days):
        difference = relativedelta(date.today(), cell_value)
        months_diff = difference.years * 12 + difference.months
        days_diff = difference.days

        if months_diff >= no_of_month:
            cell.fill = PatternFill(start_color="FF0000", fill_type="solid")  # Red BG
            cell.font = Font(color="FFFFFF", bold=True)  # White text for visibility
        elif months_diff >= (no_of_month-1) and days_diff >= no_of_days and days_diff <= (no_of_days+10):
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")  # Yellow BG
            cell.font = Font(color="000000", bold=True)  # Black text for visibility

# Usage Example
if __name__ == "__main__":
    report = Report()
    report.generate_report()
