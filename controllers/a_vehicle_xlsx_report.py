from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import os, traceback
from controllers.load_assets import *

class AVehXlsxReport:
    def __init__(self, db_to_display=None, main_heading = None, data = None, score = None):
        self.current_date = datetime.now().strftime('%d-%m-%Y')
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = os.path.join(downloads_path, f"{data['BA NO']}_Fitness_report_{timestamp}.xlsx")
        self.db_to_display = db_to_display
        self.main_header = main_heading
        self.data = data
        self.score_detail = score
        self.cell_mapping = {
            # Basic details
            "BA NO": "D8",
            "Make": "H8",
            "Type": "L8",
            "CI": "N8",
            "In Svc": "R8",  # (Adjust if needed; currently same as Type)

            # Cooling System (cells E13 to E18)
            "Fins": "E16",
            "Rad Paint": "E17",
            "Coolant": "E18",
            "CS Leakage": "E19",
            "Rad Cap": "E20",
            "Fan Belt": "E21",

            # Hyd Ramp (cells E22 to E26)
            "Hyd Oil Lvl": "E25",
            "TGS Oil Lvl": "E26",
            "Tx Oil": "E27",
            "Tx Filter": "E28",
            "Fan Mech Oil": "E29",

            # Lub Sys (cells H13 to H18)
            "Eng Oil": "H16",
            "EO Cond": "H17",
            "Oil Sump": "H18",
            "LS Leakage": "H19",
            "Oil Grade": "H20",
            "Lub": "H21",

            # Tr Sys (cells H22 to H27)
            "Tr Chain Adj": "H25",
            "Tr Chain Play": "H26",
            "Tr Pin Adj": "H27",
            "Tr Pad Thickness": "H28",
            "Sproket Wh Life": "H29",
            "Tr Tensioner": "H30",

            # Bty & Assys (cells K13 to K18)
            "Cradle Fitting": "K16",
            "Electolyte Lvl": "K17",
            "Terminals": "K18",
            "Mineral Jelly": "K19",
            "Vent Plug": "K20",
            "Bty Ser (LB)": "K21",

            # Boggy Wh (cells K22 to K24)
            "Rubber Cond": "K25",
            "Lub Pts": "K26",
            "Inner / Outer Bearing": "K27",

            # Brk Sys (cells K26 to K27)
            "Brk Fluid": "K29",
            "Brk Lever": "K30",

            # Elec Sys (cells N13 to N27)
            "Ign Sw": "N16",
            "Water Temp Guage": "N17",
            "Fuse Box": "N18",
            "Fuse Svc": "N19",
            "Oil Pressure Guage": "N20",
            "RPM Guage": "N21",
            "Oil Temp Guage": "N22",
            "Self-Starter Motor": "N23",
            "Alternator Func": "N24",
            "ES Fuel Guage": "N25",
            "Electric Harness": "N26",
            "Alternator Fan Belt": "N27",
            "Alternator Noise": "N28",
            "Horn": "N29",
            "Blower Heater": "N30",

            # Air Intake Sys (cells Q13 to Q20)
            "Air Cleaner Cond": "Q16",
            "Air Cleaner Seal": "Q17",
            "Hoses & Valves": "Q18",
            "Bluge Pump": "Q19",
            "BP Dust Cover": "Q20",
            "Hyd Oil Lvl Check": "Q21",
            "TGC Lvl Check": "Q22",
            "TGC Oil Cond": "Q23",

            # Tx Sys (cells Q23 to Q27)
            "Stall Test": "Q26",
            "Steering Planetary Gear": "Q27",
            "Final Drive Func": "Q28",
            "Tx Oil Lvl": "Q29",
            "Tx Oil Cond": "Q30",

            # Steering Con (cells T13 to T17)
            "Stick Lever Shift": "T16",
            "Stick Play": "T17",
            "Connect Rod Adj": "T18",
            "Steering Linkages": "T19",
            "Steering Pump": "T20",

            # Fuel Sys (cells T20 to T27)
            "Fuel Filter Cond": "T23",
            "Fuel Lines Leakage": "T24",
            "Fuel Filter Body": "T25",
            "Fuel Tk Strainer": "T26",
            "FS Fuel Guage": "T27",
            "Fuel Distr Cork": "T28",
            "Fuel Tk Cap": "T29",
            "Tk Inner Cond": "T30"
        }

    def generate_report(self):
        try:
            template_path = get_asset_path("assets/files/A_Veh_FItness_Check_Module.xlsx")
            # template_path = "controllers/A_Veh_FItness_Check_Module.xlsx"  # make sure this path is correct
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active  # or use wb['SheetName'] if you need a specific sheet

            # Define conditional formatting styles
            green_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            gray_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")

            thin_green_side = Side(style="thin", color="58D68D")
            thin_red_side = Side(style="thin", color="EC7063")
            thin_gray_side = Side(style="thin", color="BDC3C7")

            green_border = Border(left=thin_green_side, right=thin_green_side, top=thin_green_side, bottom=thin_green_side)
            red_border = Border(left=thin_red_side, right=thin_red_side, top=thin_red_side, bottom=thin_red_side)
            gray_border = Border(left=thin_gray_side, right=thin_gray_side, top=thin_gray_side, bottom=thin_gray_side)

            # Update the cells in the template
            for field, cell_coord in self.cell_mapping.items():
                cell = ws[cell_coord]
                value = self.data.get(field, None)
                cell.value = value  # update the cell value

                if field in ["BA NO", "Make","Type","CI","In Svc"]:
                    continue

                # Apply conditional styling based on the cell value
                if value in ["Svc", "Ok", "Up", "Complete"]:
                    cell.fill = green_fill
                    # cell.border = green_border
                elif value in ["Unsvc", "Unsatisfactory", "Down", "Incomplete"]:
                    cell.fill = red_fill
                    # cell.border = red_border
                else:
                    cell.fill = gray_fill
                    # cell.border = gray_border
            
            cell = ws["F38"]
            cell.value = f"{self.score_detail['complete_score']} / 75"

            cell = ws["K38"]
            cell.value = f"{self.score_detail['incomplete_score']}"

            cell = ws["O38"]
            cell.value = f"{self.score_detail['complete_score_percentage']}%"
            

            # Save the updated workbook as a new file
            wb.save(self.filename)
            print(f"Report saved at: {self.filename}")
            return True
        except Exception as e:
            print(f"Error While Downloading Report: {e}")
            return False